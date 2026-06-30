#!/usr/bin/env python3
# gen_projection.py —— 从 v5 Excel 模型重算三情景 → 写出 projection.json
# 依赖:python3 + openpyxl + LibreOffice(soffice)。在本地/CI 跑,不在 Railway 运行时跑。
#   pip install openpyxl
#   python3 gen_projection.py 华为数字能源经营推演模型_v5_业界稳健版.xlsx projection.json
import sys, os, json, shutil, subprocess, tempfile, datetime
import openpyxl

SOFFICE = os.environ.get("SOFFICE", "soffice")  # 或 'libreoffice'

def recalc(path):
    """用 LibreOffice 强制重算公式:置 fullCalcOnLoad 后转换一次,读回带缓存值的副本。"""
    wb = openpyxl.load_workbook(path)
    try: wb.calculation.fullCalcOnLoad = True
    except Exception: pass
    wb.save(path)
    outdir = tempfile.mkdtemp()
    subprocess.run([SOFFICE, "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", outdir, path],
                   check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    return os.path.join(outdir, os.path.basename(path))

def col(ws, r, c0=4, c1=9):  # 2025..2029 = 列 D..H
    out=[]
    for c in range(c0, c1):
        v=ws.cell(r,c).value
        out.append(round(v,1) if isinstance(v,(int,float)) else None)
    return out

def cagr(a,b,n=4):
    return round(((b/a)**(1/n)-1)*100,1) if a and b else None

def main(src, dst):
    names={1:"悲观",2:"中性",3:"乐观"}
    ns,cp,gp,seg={}, {}, {}, {}
    for scn in (1,2,3):
        tmp=os.path.join(tempfile.mkdtemp(), "m.xlsx"); shutil.copy(src, tmp)
        wb=openpyxl.load_workbook(tmp); wb["控制台"]["C3"]=scn; wb.save(tmp)
        rc=recalc(tmp)
        ws=openpyxl.load_workbook(rc, data_only=True)["数字能源总盘"]
        nm=names[scn]
        ns[nm]=[531.9,645]+col(ws,11)
        cp[nm]=[None,None]+col(ws,26)
        gp[nm]=[None,None]+col(ws,18)
        seg[nm]={"逆变器":col(ws,4),"储能":col(ws,5),"数据中心":col(ws,7),"智能电动":col(ws,8)}

    # 份额演进(中性,内生;来自份额子模型) —— 如模型份额行位置变化,在此调整
    shares=[
        {"name":"逆变器 · 欧洲","a":35,"b":30,"note":"地缘下行(高风险供应商名单)","t":"down"},
        {"name":"逆变器 · 海外其他","a":15,"b":18,"note":"新兴市场先发","t":"up"},
        {"name":"逆变器 · 中国","a":12,"b":12,"note":"格局稳固","t":"flat"},
        {"name":"储能 · 欧洲","a":10,"b":12,"note":"户储渠道+品牌","t":"up"},
        {"name":"储能 · 海外其他","a":8,"b":11,"note":"大储中标(中东)","t":"up"},
        {"name":"数据中心供电 · 全球","a":12,"b":12.2,"note":"核验上调(模块化UPS全球#1)","t":"flat"},
        {"name":"智能电动 · 中国电驱","a":14.8,"b":15.6,"note":"DriveONE 独立一供#1","t":"up"},
    ]
    out={
        "meta":{"model":"v5 业界稳健版","unit":"亿元 RMB",
                "generatedAt":datetime.date.today().isoformat(),
                "anchor2026":{"ns":875,"gp":318,"cp":131}},
        "years":[2023,2024,2025,2026,2027,2028,2029],
        "netSales":ns,"contrib":cp,"gp":gp,"seg":seg,
        "cagr":{k:cagr(ns[k][2],ns[k][6]) for k in names.values()},
        "shares":shares,
    }
    json.dump(out, open(dst,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
    print("wrote", dst)
    for k in names.values(): print(" ", k, "净销售", out["netSales"][k], "CAGR", out["cagr"][k])

if __name__=="__main__":
    src=sys.argv[1] if len(sys.argv)>1 else "华为数字能源经营推演模型_v5_业界稳健版.xlsx"
    dst=sys.argv[2] if len(sys.argv)>2 else "projection.json"
    main(src,dst)
